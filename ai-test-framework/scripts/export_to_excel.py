"""把 AI 测试数据集(yaml)导出为 Excel，便于人工快速检查。

用法:
    python export_to_excel.py                 # 导出 datasets/ 下全部 yaml
    python export_to_excel.py <a.yaml> <b.yaml> ...   # 只导出指定文件

输出:
    datasets/excel/<数据集名>.xlsx
    每个文件含 3 个 sheet:
      1. 用例明细   每条用例一行(用例ID/层/能力/维度/输入/意图/期望输出/Block/语义/标签)
      2. 覆盖矩阵   能力 x 维度 用例数(一眼看覆盖缺口)
      3. 汇总       总数/分层分布/Block 占比(公式动态统计)

依赖: openpyxl, pyyaml
"""
import argparse
import glob
import json
import os
import sys

import yaml

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少 openpyxl，请先安装: pip install openpyxl")
    sys.exit(1)

DATASETS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "..", "datasets")
OUT_DIR = os.path.join(DATASETS_DIR, "excel")

# 样式常量
HDR_FILL = PatternFill("solid", start_color="4472C4")   # 表头蓝
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial")
BASE_FONT = Font(name="Arial")
L2_FILL = PatternFill("solid", start_color="FFF2CC")     # L2 浅黄
L3_FILL = PatternFill("solid", start_color="DDEBF7")     # L3 浅蓝
BLOCK_FONT = Font(color="C00000", name="Arial")          # 拒绝用例红色
BLOCK_FILL = PatternFill("solid", start_color="FCE4EC")  # 拒绝用例浅红背景
WARN_FILL = PatternFill("solid", start_color="FFC7CE")   # 覆盖矩阵缺口红


def fmt_input(inp):
    """输入序列化为可读文本(文本直接返回，dict 转 JSON)。"""
    if isinstance(inp, dict):
        try:
            return json.dumps(inp, ensure_ascii=False, indent=1)
        except (TypeError, ValueError):
            return str(inp)
    return "" if inp is None else str(inp)


def fmt_semantic(exp):
    """期望.semantic 压缩为可读文本。"""
    sem = exp.get("semantic") or {}
    if not sem:
        return ""
    parts = []
    for k in ("contains", "fields"):
        v = sem.get(k)
        if v:
            parts.append("%s: %s" % (k, " | ".join(str(x) for x in v)))
    if sem.get("any_of"):
        parts.append("any_of=true")
    return "  ".join(parts)


def fmt_params(exp):
    params = exp.get("params")
    if not params:
        return ""
    return json.dumps(params, ensure_ascii=False) if isinstance(params, dict) \
        else str(params)


def sheet_details(ws, cases):
    """用例明细 sheet。"""
    headers = ["用例ID", "层", "能力", "维度", "输入", "意图",
               "参数", "期望输出", "Block", "语义(contains/fields)", "标签"]
    ws.append(headers)
    for c in cases:
        exp = c.get("期望") or {}
        ws.append([
            c.get("用例ID", ""),
            c.get("层", ""),
            c.get("能力", ""),
            c.get("维度", ""),
            fmt_input(c.get("输入")),
            exp.get("intent", ""),
            fmt_params(exp),
            exp.get("output", ""),
            "是" if exp.get("block") else "否",
            fmt_semantic(exp),
            "、".join(c.get("标签") or []),
        ])
    # 表头样式 + 冻结 + 筛选
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)),
                                      max(2, ws.max_row))
    # 层着色 + block 标红 + 列宽
    for r in range(2, ws.max_row + 1):
        layer = ws.cell(row=r, column=2).value
        if layer == "L2":
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = L2_FILL
        elif layer == "L3":
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = L3_FILL
        if ws.cell(row=r, column=9).value == "是":
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=col)
                cell.font = BLOCK_FONT
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = BLOCK_FILL
    for col, w in enumerate((12, 6, 18, 16, 52, 26, 26, 34, 8, 46, 14), 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22


def sheet_matrix(ws, cases, abilities):
    """能力 x 维度 覆盖矩阵(数字=用例数，0 标红)。"""
    dims = []
    for c in cases:
        d = c.get("维度")
        if d and d not in dims:
            dims.append(d)
    caps = []
    for c in cases:
        a = c.get("能力")
        if a and a not in caps:
            caps.append(a)
    caps.sort()
    # 首行维度
    ws.cell(row=1, column=1, value="能力 \\ 维度")
    for j, d in enumerate(dims, 2):
        cell = ws.cell(row=1, column=j, value=d)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell = ws.cell(row=1, column=len(dims) + 2, value="合计")
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center")
    cnt = {}
    for c in cases:
        key = (c.get("能力"), c.get("维度"))
        cnt[key] = cnt.get(key, 0) + 1
    for i, a in enumerate(caps, 2):
        ws.cell(row=i, column=1, value=a).font = BASE_FONT
        total = 0
        for j, d in enumerate(dims, 2):
            n = cnt.get((a, d), 0)
            total += n
            cell = ws.cell(row=i, column=j, value=n)
            cell.alignment = Alignment(horizontal="center")
            if n == 0:
                cell.fill = WARN_FILL
        cell = ws.cell(row=i, column=len(dims) + 2, value=total)
        cell.font = Font(bold=True, name="Arial")
        cell.alignment = Alignment(horizontal="center")
    # 表头样式
    ws.cell(row=1, column=1).fill = HDR_FILL
    ws.cell(row=1, column=1).font = HDR_FONT
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 22
    for j in range(2, len(dims) + 3):
        ws.column_dimensions[get_column_letter(j)].width = 14


def sheet_summary(ws, cases, n_rows):
    """汇总 sheet：总数/分层分布/Block 占比(用 COUNTIF 公式动态统计)。"""
    last = n_rows + 1
    ws.append(["指标", "值", "说明"])
    ws.append(["用例总数", len(cases), "全部用例数"])
    ws.append(["L1 用例数", "=COUNTIF(用例明细!$B$2:$B$%d,\"L1\")" % last,
               "=IF($B$2=0,\"\",ROUND(B3/$B$2,3))"])
    ws.append(["L2 用例数", "=COUNTIF(用例明细!$B$2:$B$%d,\"L2\")" % last,
               "=IF($B$2=0,\"\",ROUND(B4/$B$2,3))"])
    ws.append(["L3 用例数", "=COUNTIF(用例明细!$B$2:$B$%d,\"L3\")" % last,
               "=IF($B$2=0,\"\",ROUND(B5/$B$2,3))"])
    ws.append(["Block(拒绝) 用例数",
               "=COUNTIF(用例明细!$I$2:$I$%d,\"是\")" % last,
               "期望 block=true 的用例数"])
    ws.append(["Block 占比", "=IF($B$2=0,\"\",ROUND(B6/$B$2,3))",
               "拒绝类用例占总用例比例"])
    for row in range(1, ws.max_row + 1):
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            if row == 1:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
            else:
                cell.font = BASE_FONT
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 40


def export_one(path):
    with open(path, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    cases = data.get("用例列表") or []
    name = os.path.splitext(os.path.basename(path))[0]
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "用例明细"
    sheet_details(ws1, cases)
    ws2 = wb.create_sheet("覆盖矩阵")
    sheet_matrix(ws2, cases, [])
    ws3 = wb.create_sheet("汇总")
    sheet_summary(ws3, cases, len(cases))
    out = os.path.join(OUT_DIR, name + ".xlsx")
    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(out)
    return out, len(cases)


def main():
    parser = argparse.ArgumentParser(description="AI 测试数据集 yaml -> Excel")
    parser.add_argument("files", nargs="*", help="数据集 yaml 路径(缺省=datasets/ 下全部)")
    args = parser.parse_args()
    if args.files:
        files = [f for f in args.files if os.path.isfile(f)]
    else:
        files = sorted(glob.glob(os.path.join(DATASETS_DIR, "*.yaml")))
    if not files:
        print("未找到任何数据集 yaml")
        sys.exit(1)
    for path in files:
        try:
            out, n = export_one(path)
            print("OK  %s  (%d 条 -> %s)" % (os.path.basename(path), n, out))
        except Exception as e:  # noqa: BLE001
            print("FAIL %s: %s" % (os.path.basename(path), e))


if __name__ == "__main__":
    main()
